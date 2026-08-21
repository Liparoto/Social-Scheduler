"""End-to-end and safety tests for the export entry point."""

from __future__ import annotations

import hashlib
import json
import sqlite3
import zipfile
from datetime import datetime, timezone

import pytest
from openpyxl import load_workbook

from worker.config import Config
from worker.export.__main__ import DB_COPY_NAME, copy_database, main, open_readonly, run_export

NOW = datetime(2026, 7, 24, 19, 30, tzinfo=timezone.utc)


def _seed_assets(config, conn):
    """Give every asset row a real file on disk, so copying succeeds."""
    config.asset_storage_dir.mkdir(parents=True, exist_ok=True)
    for row in conn.execute("SELECT id, storage_path FROM assets"):
        path = config.asset_storage_dir / row["storage_path"].split("/")[-1]
        path.write_bytes(f"bytes-{row['id']}".encode())
        conn.execute(
            "UPDATE assets SET storage_path = ? WHERE id = ?", (path.name, row["id"])
        )
    conn.commit()


def test_open_readonly_rejects_writes(db_path):
    conn = open_readonly(db_path)

    with pytest.raises(sqlite3.OperationalError) as excinfo:
        conn.execute("INSERT INTO tags (name) VALUES ('nope')")

    # A missing table would raise the same exception type, so pin down the message
    # too — this confirms query_only is what blocked it, not a schema mistake.
    assert "readonly" in str(excinfo.value).lower()

    conn.close()


def test_export_leaves_the_database_byte_identical(config, conn, make_publication, tmp_path):
    make_publication()
    _seed_assets(config, conn)
    conn.close()
    before = hashlib.sha256(config.database_path.read_bytes()).hexdigest()

    run_export(config, out_root=tmp_path / "out", now=NOW)

    assert hashlib.sha256(config.database_path.read_bytes()).hexdigest() == before


def test_export_writes_no_token_into_any_file(config, conn, make_publication, tmp_path):
    make_publication(with_token=True)
    _seed_assets(config, conn)
    conn.close()

    out_dir, _ = run_export(config, out_root=tmp_path / "out", now=NOW)

    # The token VALUE must not appear anywhere, in any file, ever. No exceptions.
    # The column NAME is a separate matter, handled below.
    secret_value = b"tok-123"
    column_name = b"access_token"

    def _assert_clean(blob: bytes, where: str, needles) -> None:
        for needle in needles:
            assert needle not in blob, f"token leaked into {where}"

    for path in out_dir.rglob("*"):
        if not path.is_file():
            continue
        # The database copy is the one file that legitimately contains the STRING
        # "access_token": its schema declares the column, in the CREATE TABLE text
        # SQLite stores in sqlite_master. That is a column name, not a credential, and
        # it cannot be removed without shipping a different schema than the one the app
        # reads — which would defeat the copy's only purpose, being restorable.
        #
        # For every OTHER file the column name staying forbidden is the point: those are
        # built by naming what to emit, so "access_token" appearing in the spreadsheet or
        # the JSON means a credential column reached the allow-list.
        #
        # The value check below is unconditional, and the copy gets a stronger assertion
        # than a byte scan afterwards.
        needles = (secret_value,) if path.name == DB_COPY_NAME else (secret_value, column_name)
        # openpyxl DEFLATE-compresses every member of an .xlsx (it's a ZIP archive),
        # so a raw byte scan of the file never sees a token that IS present in a
        # cell — it only sees compressed bytes. Scan each member's decompressed
        # bytes instead for any zip archive; everything else gets the raw scan.
        if zipfile.is_zipfile(path):
            with zipfile.ZipFile(path) as archive:
                for member in archive.namelist():
                    _assert_clean(archive.read(member), f"{path.name}!{member}", needles)
        else:
            _assert_clean(path.read_bytes(), path.name, needles)


def test_the_database_copy_is_scrubbed_of_every_credential(
    config, conn, make_publication, tmp_path
):
    """The byte scan proves the token text is gone; this proves the COLUMN is empty.

    A byte scan can be satisfied by accident — compression, or a value that happens not to
    appear literally. Reading the column back through SQLite cannot.
    """
    make_publication(with_token=True)
    _seed_assets(config, conn)
    conn.close()

    out_dir, _ = run_export(config, out_root=tmp_path / "out", now=NOW)

    copy = sqlite3.connect(str(out_dir / DB_COPY_NAME))
    try:
        # Present and readable — this is a real, restorable database, not a stripped one.
        assert copy.execute("SELECT COUNT(*) FROM channels").fetchone()[0] > 0
        for column in ("access_token", "token_expires_at"):
            values = [
                row[0]
                for row in copy.execute(f"SELECT {column} FROM channels")
                if row[0] not in (None, "")
            ]
            assert values == [], f"{column} survived into the exported copy"
    finally:
        copy.close()


def test_the_scrub_finds_credential_columns_by_shape_not_by_name(config, conn, tmp_path):
    """Adding a new credential column must not silently start leaking.

    The allow-list that protects the rest of the export is held to this same structural
    rule. A copy of the whole database cannot choose what it reads, so it has to recognise a
    credential by the shape of its name and empty it by default.
    """
    conn.execute("ALTER TABLE channels ADD COLUMN refresh_secret TEXT")
    conn.execute(
        "INSERT INTO channels (platform, account_name, refresh_secret)"
        " VALUES ('instagram', 'future', 'super-secret-value')"
    )
    conn.commit()
    conn.close()

    out_dir = tmp_path / "out"
    out_dir.mkdir(parents=True)
    dest = copy_database(config.database_path, out_dir)

    assert b"super-secret-value" not in dest.read_bytes()
    copy = sqlite3.connect(str(dest))
    try:
        assert copy.execute(
            "SELECT COUNT(*) FROM channels WHERE refresh_secret IS NOT NULL"
        ).fetchone()[0] == 0
    finally:
        copy.close()


def test_main_reports_a_corrupt_database_instead_of_crashing(config, tmp_path, monkeypatch, capsys):
    # A file that exists but is not a valid SQLite database — e.g. a truncated or
    # otherwise corrupted DB file. sqlite3 raises DatabaseError for this, which does
    # NOT subclass OSError, so it must be caught explicitly in main().
    bad_db = tmp_path / "corrupt.db"
    bad_db.write_bytes(b"this is not a database")
    bad_config = Config(**{**config.__dict__, "database_path": bad_db})

    monkeypatch.setattr(Config, "from_env", staticmethod(lambda: bad_config))
    # Don't let a failed run scribble a folder into the real ~/Documents.
    monkeypatch.setattr("worker.export.__main__.DEFAULT_OUT_ROOT", tmp_path / "out")

    exit_code = main([])

    assert exit_code == 1
    out = capsys.readouterr().out
    assert "not modified" in out.lower()


def test_main_returns_1_on_an_unexpected_exception_instead_of_crashing(
    config, tmp_path, monkeypatch, capsys
):
    # Anything not OSError/sqlite3.Error (e.g. openpyxl's IllegalCharacterError, or
    # any other bug) must still exit cleanly for a non-technical person who
    # double-clicked an icon — never a raw traceback.
    monkeypatch.setattr(Config, "from_env", staticmethod(lambda: config))
    monkeypatch.setattr("worker.export.__main__.DEFAULT_OUT_ROOT", tmp_path / "out")

    def _boom(*args, **kwargs):
        raise RuntimeError("kaboom")

    monkeypatch.setattr("worker.export.__main__.collect_all", _boom)

    exit_code = main([])

    assert exit_code == 1
    out = capsys.readouterr().out
    assert "not modified" in out.lower()


def test_export_creates_the_expected_folder_layout(config, conn, make_publication, tmp_path):
    make_publication()
    _seed_assets(config, conn)
    conn.close()

    out_dir, result = run_export(config, out_root=tmp_path / "out", now=NOW)

    assert out_dir.name == "2026-07-24-1930"
    assert (out_dir / "README.txt").is_file()
    assert (out_dir / "SocialScheduler-Export.xlsx").is_file()
    assert (out_dir / "export.json").is_file()
    assert result.copied == 1
    assert (out_dir / "images").is_dir()


def test_export_never_overwrites_a_previous_run(config, conn, make_publication, tmp_path):
    make_publication()
    _seed_assets(config, conn)
    conn.close()
    out_root = tmp_path / "out"

    first, _ = run_export(config, out_root=out_root, now=NOW)
    (first / "sentinel.txt").write_text("do not clobber me")
    second, _ = run_export(config, out_root=out_root, now=NOW)

    assert second != first
    assert (first / "sentinel.txt").read_text() == "do not clobber me"


def test_export_of_an_empty_database_still_produces_a_valid_workbook(config, tmp_path):
    out_dir, result = run_export(config, out_root=tmp_path / "out", now=NOW)

    book = load_workbook(out_dir / "SocialScheduler-Export.xlsx")

    assert book.sheetnames == [
        "Posts", "Sends", "Metrics", "Assets", "Channels", "Channel groups",
    ]
    assert result.copied == 0
    assert json.loads((out_dir / "export.json").read_text())["posts"] == []


def test_export_completes_when_an_asset_file_is_missing(config, conn, make_publication, tmp_path):
    make_publication()  # asset rows exist, but no files were written to disk
    conn.close()

    out_dir, result = run_export(config, out_root=tmp_path / "out", now=NOW)

    assert result.missing_asset_ids
    assert "could not be found" in (out_dir / "README.txt").read_text()
    assert (out_dir / "SocialScheduler-Export.xlsx").is_file()


def test_export_holds_one_snapshot_so_a_mid_collection_commit_is_not_visible(
    config, conn, make_publication, tmp_path, monkeypatch
):
    # sqlite3 does not open a transaction for plain reads: without one, each of
    # collect_all's 6+ SELECTs takes its own WAL snapshot, and a commit from
    # another connection (e.g. the worker daemon) between two of them can leave
    # the export self-inconsistent (a Sends row whose post doesn't exist, etc.).
    # We interleave a write from a SEPARATE connection after the export's first
    # read (which establishes its snapshot) but before its later reads, and
    # confirm the later read does not see it.
    pub = make_publication()
    post_id, channel_id = pub["post_id"], pub["channel_id"]
    conn.commit()
    conn.close()

    import worker.export.collect as collect_mod

    original_add_rollups = collect_mod.add_rollups

    def _add_rollups_with_interleaved_write(c, posts):
        other = sqlite3.connect(str(config.database_path))
        other.execute(
            "INSERT INTO publications (post_id, channel_id, scheduled_at, status)"
            " VALUES (?, ?, '2026-08-01T00:00:00+00:00', 'scheduled')",
            (post_id, channel_id),
        )
        other.commit()
        other.close()
        return original_add_rollups(c, posts)

    monkeypatch.setattr(collect_mod, "add_rollups", _add_rollups_with_interleaved_write)

    out_dir, _ = run_export(config, out_root=tmp_path / "out", now=NOW)

    data = json.loads((out_dir / "export.json").read_text())
    # Only the original publication (from make_publication) may appear — the one
    # inserted by the separate connection, mid-collection, must not leak in.
    assert len(data["sends"]) == 1


def test_export_transaction_is_open_during_collection(
    config, conn, make_publication, tmp_path, monkeypatch
):
    make_publication()
    conn.commit()
    conn.close()

    import worker.export.collect as collect_mod

    seen_in_transaction = {}
    original_collect_all = collect_mod.collect_all

    def _spy(c, generated_at):
        seen_in_transaction["value"] = c.in_transaction
        return original_collect_all(c, generated_at)

    monkeypatch.setattr("worker.export.__main__.collect_all", _spy)

    run_export(config, out_root=tmp_path / "out", now=NOW)

    assert seen_in_transaction["value"] is True


def test_the_database_copy_captures_rows_still_sitting_in_the_wal(config, conn, tmp_path):
    """Why the backup API, and not `cp`.

    In WAL mode a committed row can live entirely in the -wal sidecar, with the main .db
    file not yet containing it. Copying the file alone would produce a backup that silently
    lacks recent work — and, worse, can catch a torn page mid-checkpoint. The backup API
    reads through the engine, so it sees the committed state including the WAL.
    """
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("INSERT INTO tags (name) VALUES ('only-in-the-wal')")
    conn.commit()
    # Deliberately NOT checkpointed and NOT closed: the row is committed, and this is
    # exactly the state the live install is in whenever the worker is running.

    out_dir = tmp_path / "out"
    out_dir.mkdir(parents=True)
    dest = copy_database(config.database_path, out_dir)

    copy = sqlite3.connect(str(dest))
    try:
        names = [r[0] for r in copy.execute("SELECT name FROM tags")]
        assert "only-in-the-wal" in names
    finally:
        copy.close()
    conn.close()


def test_the_export_still_writes_the_copy_when_images_are_missing(
    config, conn, make_publication, tmp_path
):
    """The copy is written FIRST, so the long tail of things that can fail cannot cost it.

    No files are seeded here, so every image copy fails — which is the realistic partial
    failure (a file deleted from the asset store behind the app's back).
    """
    make_publication()
    conn.close()

    out_dir, copy_result = run_export(config, out_root=tmp_path / "out", now=NOW)

    assert copy_result.missing_asset_ids, "precondition: images really did fail to copy"
    assert (out_dir / DB_COPY_NAME).is_file(), "a partial export still restores"
