"""Tests for turning an ExportBundle into files on disk."""

from __future__ import annotations

from worker.export.collect import ExportBundle, ExportedPost, PostImage
from worker.export.write import copy_images

GENERATED_AT = "2026-07-24T19:30:00+00:00"


def _bundle(posts):
    return ExportBundle(
        generated_at=GENERATED_AT, posts=posts, sends=[], metrics=[],
        assets=[], channels=[],
    )


def _post(post_id, images):
    return ExportedPost(
        post_id=post_id, caption="Test Post", first_comment=None, post_type="single",
        content_kind="evergreen", content_status="ready", status="draft",
        cooldown_days=None, created_by=None, created_at=None, updated_at=None,
        images=images,
    )


def _image(asset_id, name, storage_path, publish_path=None, published_name=None):
    return PostImage(
        asset_id=asset_id, sort_order=0, export_filename=name,
        published_filename=published_name, storage_path=storage_path,
        publish_path=publish_path,
    )


def test_copy_images_writes_originals_into_the_images_folder(tmp_path):
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "a.jpg").write_bytes(b"original-bytes")
    out = tmp_path / "out"

    result = copy_images(
        _bundle([_post(42, [_image(1, "0042_test-post_1.jpg", "a.jpg")])]),
        asset_root=assets,
        out_dir=out,
    )

    assert result.copied == 1
    assert result.missing_asset_ids == set()
    assert (out / "images" / "0042_test-post_1.jpg").read_bytes() == b"original-bytes"


def test_copy_images_writes_conformed_copies_into_a_separate_folder(tmp_path):
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "a.jpg").write_bytes(b"original-bytes")
    (assets / "a-pub.jpg").write_bytes(b"cropped-bytes")
    out = tmp_path / "out"

    copy_images(
        _bundle([_post(42, [_image(
            1, "0042_test-post_1.jpg", "a.jpg",
            publish_path="a-pub.jpg", published_name="0042_test-post_1.jpg",
        )])]),
        asset_root=assets,
        out_dir=out,
    )

    assert (out / "images" / "0042_test-post_1.jpg").read_bytes() == b"original-bytes"
    assert (out / "images-published" / "0042_test-post_1.jpg").read_bytes() == b"cropped-bytes"


def test_copy_images_skips_the_published_folder_when_nothing_was_conformed(tmp_path):
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "a.jpg").write_bytes(b"x")
    out = tmp_path / "out"

    copy_images(
        _bundle([_post(42, [_image(1, "0042_test-post_1.jpg", "a.jpg")])]),
        asset_root=assets,
        out_dir=out,
    )

    assert not (out / "images-published").exists()


def test_copy_images_records_missing_files_instead_of_raising(tmp_path):
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "present.jpg").write_bytes(b"here")
    out = tmp_path / "out"

    result = copy_images(
        _bundle([_post(42, [
            _image(1, "0042_test-post_1.jpg", "present.jpg"),
            _image(2, "0042_test-post_2.jpg", "gone.jpg"),
        ])]),
        asset_root=assets,
        out_dir=out,
    )

    # A partial backup you know is partial beats a crash and no backup.
    assert result.copied == 1
    assert result.missing_asset_ids == {2}
    assert len(result.problems) == 1
    assert "gone.jpg" in result.problems[0]


def test_copy_images_exports_a_shared_asset_once_per_post(tmp_path):
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "shared.jpg").write_bytes(b"shared-bytes")
    out = tmp_path / "out"

    result = copy_images(
        _bundle([
            _post(42, [_image(1, "0042_test-post_1.jpg", "shared.jpg")]),
            _post(51, [_image(1, "0051_test-post_1.jpg", "shared.jpg")]),
        ]),
        asset_root=assets,
        out_dir=out,
    )

    assert result.copied == 2
    assert (out / "images" / "0042_test-post_1.jpg").exists()
    assert (out / "images" / "0051_test-post_1.jpg").exists()


def test_copy_images_handles_a_bundle_with_no_images(tmp_path):
    out = tmp_path / "out"

    result = copy_images(_bundle([_post(42, [])]), asset_root=tmp_path, out_dir=out)

    assert result.copied == 0
    assert result.missing_asset_ids == set()


def test_copy_images_missing_conformed_copy_does_not_mark_the_original_missing(tmp_path):
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "a.jpg").write_bytes(b"original-bytes")
    out = tmp_path / "out"

    result = copy_images(
        _bundle([_post(42, [_image(
            1, "0042_test-post_1.jpg", "a.jpg",
            publish_path="gone-pub.jpg", published_name="0042_test-post_1.jpg",
        )])]),
        asset_root=assets,
        out_dir=out,
    )

    # The original exported fine — it must never be reported as missing just
    # because the conformed (Instagram-ready) copy wasn't there.
    assert result.copied == 1
    assert result.missing_asset_ids == set()
    assert len(result.problems) == 1
    assert "gone-pub.jpg" in result.problems[0]


def test_copy_images_missing_original_marks_the_asset_missing_even_if_conformed_copy_exists(tmp_path):
    assets = tmp_path / "assets"
    assets.mkdir()
    (assets / "a-pub.jpg").write_bytes(b"cropped-bytes")
    out = tmp_path / "out"

    result = copy_images(
        _bundle([_post(42, [_image(
            1, "0042_test-post_1.jpg", "gone.jpg",
            publish_path="a-pub.jpg", published_name="0042_test-post_1.jpg",
        )])]),
        asset_root=assets,
        out_dir=out,
    )

    assert result.missing_asset_ids == {1}


import json

from worker.export.collect import ExportedChannel, ExportedSend
from worker.export.write import write_json


def _channel():
    return ExportedChannel(
        channel_id=1, platform="instagram", account_name="Test IG",
        business_label=None, timezone="UTC", is_active=True, requires_approval=False,
        autofill_enabled=False, cadence_config=None, min_queue_depth=0,
        target_queue_depth=0, reuse_min_age_days=180, remote_account_id="178414",
        linked_page_id=None,
    )


def test_write_json_round_trips_the_bundle(tmp_path):
    bundle = _bundle([_post(42, [_image(1, "0042_test-post_1.jpg", "assets/a.jpg")])])
    bundle.channels = [_channel()]

    path = write_json(bundle, tmp_path)
    data = json.loads(path.read_text())

    assert path.name == "export.json"
    assert data["generated_at"] == GENERATED_AT
    assert data["posts"][0]["post_id"] == 42
    assert data["posts"][0]["images"][0]["export_filename"] == "0042_test-post_1.jpg"
    assert data["channels"][0]["account_name"] == "Test IG"


def test_write_json_contains_no_token_field_anywhere(tmp_path):
    bundle = _bundle([_post(42, [])])
    bundle.channels = [_channel()]

    raw = write_json(bundle, tmp_path).read_text()

    assert "access_token" not in raw
    assert "token_expires_at" not in raw


def test_write_json_preserves_non_ascii_captions(tmp_path):
    post = _post(42, [])
    post.caption = "Café ☕ mobility"
    bundle = _bundle([post])

    data = json.loads(write_json(bundle, tmp_path).read_text())

    assert data["posts"][0]["caption"] == "Café ☕ mobility"


def test_write_json_includes_sends_and_records_schema_version(tmp_path):
    bundle = _bundle([])
    bundle.sends = [ExportedSend(
        publication_id=7, post_id=42, caption_preview="hello", channel_label="Test IG (instagram)",
        scheduled_at_utc="2026-07-24T18:00:00+00:00", scheduled_at_local="2026-07-24 18:00",
        published_at_utc=None, published_at_local=None, status="scheduled", is_held=False,
        is_dry_run=False, attempt_count=0, last_error=None, remote_post_id=None,
    )]

    data = json.loads(write_json(bundle, tmp_path).read_text())

    assert data["sends"][0]["publication_id"] == 7
    assert data["format_version"] == 1


from openpyxl import load_workbook

from worker.export.collect import ExportedAsset, ExportedMetric
from worker.export.write import write_workbook


def _rows(sheet):
    """Sheet contents as a list of dicts keyed by the header row."""
    values = list(sheet.values)
    header = values[0]
    return [dict(zip(header, row)) for row in values[1:]]


def test_write_workbook_creates_all_five_tabs(tmp_path):
    path = write_workbook(_bundle([]), tmp_path, missing_asset_ids=set())
    book = load_workbook(path)

    assert path.name == "SocialScheduler-Export.xlsx"
    assert book.sheetnames == ["Posts", "Sends", "Metrics", "Assets", "Channels"]


def test_write_workbook_writes_headers_even_for_an_empty_database(tmp_path):
    book = load_workbook(write_workbook(_bundle([]), tmp_path, missing_asset_ids=set()))

    assert next(book["Posts"].values)[0] == "post_id"
    assert book["Posts"].max_row == 1


def test_posts_tab_joins_multi_value_fields_with_commas(tmp_path):
    post = _post(42, [
        _image(1, "0042_test-post_1.jpg", "assets/a.jpg"),
        _image(2, "0042_test-post_2.jpg", "assets/b.jpg"),
    ])
    post.tags = ["mobility", "balance"]
    post.target_channels = ["Test IG"]
    post.times_posted = 3
    post.total_reach = 900

    book = load_workbook(write_workbook(_bundle([post]), tmp_path, missing_asset_ids=set()))
    row = _rows(book["Posts"])[0]

    assert row["tags"] == "mobility, balance"
    assert row["image_files"] == "0042_test-post_1.jpg, 0042_test-post_2.jpg"
    assert row["times_posted"] == 3
    assert row["total_reach"] == 900


def test_channels_tab_has_no_token_column(tmp_path):
    bundle = _bundle([])
    bundle.channels = [_channel()]

    book = load_workbook(write_workbook(bundle, tmp_path, missing_asset_ids=set()))
    header = next(book["Channels"].values)

    assert "access_token" not in header
    assert "token_expires_at" not in header
    assert "account_name" in header


def test_assets_tab_flags_files_that_were_missing_from_disk(tmp_path):
    bundle = _bundle([_post(42, [_image(1, "0042_test-post_1.jpg", "assets/gone.jpg")])])
    bundle.assets = [ExportedAsset(
        asset_id=1, content_hash="h", media_kind="image", original_filename="gone.jpg",
        storage_path="assets/gone.jpg", publish_path=None, conform_mode="none",
        needs_review=False, mime_type="image/jpeg", width=1080, height=1080, byte_size=None,
    )]

    book = load_workbook(write_workbook(bundle, tmp_path, missing_asset_ids={1}))
    row = _rows(book["Assets"])[0]

    assert row["exported_filename"] == "MISSING"
    assert row["used_by_posts"] == "42"


def test_assets_tab_lists_every_post_that_uses_a_shared_asset(tmp_path):
    bundle = _bundle([
        _post(42, [_image(1, "0042_test-post_1.jpg", "assets/s.jpg")]),
        _post(51, [_image(1, "0051_test-post_1.jpg", "assets/s.jpg")]),
    ])
    bundle.assets = [ExportedAsset(
        asset_id=1, content_hash="h", media_kind="image", original_filename="s.jpg",
        storage_path="assets/s.jpg", publish_path=None, conform_mode="none",
        needs_review=False, mime_type="image/jpeg", width=1080, height=1080, byte_size=10,
    )]

    row = _rows(load_workbook(
        write_workbook(bundle, tmp_path, missing_asset_ids=set())
    )["Assets"])[0]

    assert row["used_by_posts"] == "42, 51"
    assert row["exported_filename"] == "0042_test-post_1.jpg, 0051_test-post_1.jpg"


def test_metrics_tab_omits_raw_json_but_keeps_every_snapshot(tmp_path):
    bundle = _bundle([])
    bundle.metrics = [
        ExportedMetric(
            publication_id=7, post_id=42, fetched_at=f"2026-07-0{n}T00:00:00+00:00",
            reach=n * 100, impressions=None, likes=n, comments=None, saves=None,
            shares=None, video_views=None, raw_json='{"big":"payload"}',
        )
        for n in (1, 5)
    ]

    sheet = load_workbook(write_workbook(bundle, tmp_path, missing_asset_ids=set()))["Metrics"]

    assert "raw_json" not in next(sheet.values)
    assert len(_rows(sheet)) == 2


def test_assets_tab_keeps_published_copy_filename_index_aligned_with_used_by_posts(tmp_path):
    # Asset used by three posts; only the 2nd and 3rd were conformed. Without
    # index-alignment, published_copy_filename would only have two entries and
    # a reader would misread post 42 as having produced the first conformed file.
    bundle = _bundle([
        _post(42, [_image(1, "0042_test-post_1.jpg", "assets/s.jpg")]),
        _post(51, [_image(
            1, "0051_test-post_1.jpg", "assets/s.jpg",
            publish_path="assets/s-pub-51.jpg", published_name="0051_test-post_1.jpg",
        )]),
        _post(60, [_image(
            1, "0060_test-post_1.jpg", "assets/s.jpg",
            publish_path="assets/s-pub-60.jpg", published_name="0060_test-post_1.jpg",
        )]),
    ])
    bundle.assets = [ExportedAsset(
        asset_id=1, content_hash="h", media_kind="image", original_filename="s.jpg",
        storage_path="assets/s.jpg", publish_path=None, conform_mode="none",
        needs_review=False, mime_type="image/jpeg", width=1080, height=1080, byte_size=10,
    )]

    row = _rows(load_workbook(
        write_workbook(bundle, tmp_path, missing_asset_ids=set())
    )["Assets"])[0]

    used_by = row["used_by_posts"].split(", ")
    published = row["published_copy_filename"].split(", ")

    assert used_by == ["42", "51", "60"]
    assert len(published) == len(used_by)
    assert published[0] == "-"
    assert published[1] == "0051_test-post_1.jpg"
    assert published[2] == "0060_test-post_1.jpg"


def test_assets_tab_published_copy_filename_is_empty_when_nothing_was_conformed(tmp_path):
    bundle = _bundle([
        _post(42, [_image(1, "0042_test-post_1.jpg", "assets/s.jpg")]),
        _post(51, [_image(1, "0051_test-post_1.jpg", "assets/s.jpg")]),
    ])
    bundle.assets = [ExportedAsset(
        asset_id=1, content_hash="h", media_kind="image", original_filename="s.jpg",
        storage_path="assets/s.jpg", publish_path=None, conform_mode="none",
        needs_review=False, mime_type="image/jpeg", width=1080, height=1080, byte_size=10,
    )]

    row = _rows(load_workbook(
        write_workbook(bundle, tmp_path, missing_asset_ids=set())
    )["Assets"])[0]

    assert row["published_copy_filename"] in (None, "")


def test_sends_tab_has_published_at_utc_adjacent_to_published_at_local(tmp_path):
    bundle = _bundle([])
    bundle.sends = [ExportedSend(
        publication_id=7, post_id=42, caption_preview="hello", channel_label="Test IG (instagram)",
        scheduled_at_utc="2026-07-24T18:00:00+00:00", scheduled_at_local="2026-07-24 18:00",
        published_at_utc="2026-07-24T18:05:00+00:00", published_at_local="2026-07-24 18:05",
        status="published", is_held=False, is_dry_run=False, attempt_count=1,
        last_error=None, remote_post_id="123456",
    )]

    book = load_workbook(write_workbook(bundle, tmp_path, missing_asset_ids=set()))
    header = list(next(book["Sends"].values))

    local_idx = header.index("published_at_local")
    assert header[local_idx + 1] == "published_at_utc"

    row = _rows(book["Sends"])[0]
    assert row["published_at_utc"] == "2026-07-24T18:05:00+00:00"


from worker.export.write import CopyResult, write_readme


def test_readme_summarizes_the_contents(tmp_path):
    bundle = _bundle([_post(42, [])])
    bundle.channels = [_channel()]

    text = write_readme(bundle, tmp_path, CopyResult(copied=3)).read_text()

    assert "1 post" in text
    assert "3 image file" in text
    assert "SocialScheduler-Export.xlsx" in text


def test_readme_reports_missing_files_prominently(tmp_path):
    result = CopyResult(copied=1, missing_asset_ids={2, 3},
                        problems=["asset 2: file not found at /x/gone.jpg"])

    text = write_readme(_bundle([]), tmp_path, result).read_text()

    assert "2 image file(s) could not be found" in text
    assert "gone.jpg" in text


def test_readme_says_so_when_nothing_was_missing(tmp_path):
    text = write_readme(_bundle([]), tmp_path, CopyResult(copied=1)).read_text()

    assert "could not be found" not in text
    assert "No problems" in text


def test_readme_never_mentions_tokens(tmp_path):
    bundle = _bundle([])
    bundle.channels = [_channel()]

    text = write_readme(bundle, tmp_path, CopyResult()).read_text()

    assert "access_token" not in text
