"""Turn an ExportBundle into the files a human (and a future importer) can use.

Everything here writes only inside the run's own output directory. Nothing in this
module touches the database or the network.
"""

from __future__ import annotations

import json
import shutil
from dataclasses import dataclass, field, asdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from worker.export.collect import ExportBundle, ExportedAsset, slugify

IMAGES_DIR = "images"
PUBLISHED_DIR = "images-published"
UNLINKED_DIR = "images-unlinked"


@dataclass
class CopyResult:
    copied: int = 0
    # Means exactly one thing: the asset's ORIGINAL (storage_path) could not be
    # exported. A missing/unreadable CONFORMED copy is recorded in `problems`
    # only — never here — since the irreplaceable source still exported fine.
    missing_asset_ids: set[int] = field(default_factory=set)
    problems: list[str] = field(default_factory=list)


def _resolve(asset_root: Path, stored_path: str) -> Path:
    """Where an asset row's file actually lives.

    storage_path holds a bare content-hash filename relative to the asset store
    (verified against the live database). Absolute paths are tolerated in case an
    install ever stores them that way.
    """
    candidate = Path(stored_path)
    return candidate if candidate.is_absolute() else asset_root / candidate


def _copy_one(
    src: Path, dest_dir: Path, name: str, result: CopyResult, asset_id: int, *, is_original: bool
) -> bool:
    kind = "original" if is_original else "conformed copy"
    if not src.is_file():
        if is_original:
            result.missing_asset_ids.add(asset_id)
        result.problems.append(f"asset {asset_id}: {kind} not found at {src}")
        return False
    try:
        dest_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dest_dir / name)
    except OSError as exc:
        if is_original:
            result.missing_asset_ids.add(asset_id)
        result.problems.append(f"asset {asset_id}: could not copy {kind} {src} ({exc})")
        return False
    result.copied += 1
    return True


def unlinked_filename(asset: ExportedAsset) -> str:
    """Filename for an asset backed up because it belongs to no post.

    Mirrors collect.py's `_images_for` naming: an id-prefixed, slugified name so
    it sorts and reads the same way as post-linked images do.
    """
    ext = Path(asset.storage_path).suffix.lower() or ".bin"
    # Slugify only the stem (filename without extension) so "Orphan Shot.jpg"
    # becomes "orphan-shot", not "orphan-shot-jpg".
    stem = Path(asset.original_filename).stem if asset.original_filename else None
    return f"{asset.asset_id:04d}_{slugify(stem)}{ext}"


def copy_images(bundle: ExportBundle, asset_root: Path, out_dir: Path) -> CopyResult:
    """Copy every post's images under their exported names, plus any asset that
    is not linked to a post at all.

    An asset shared by two posts is written once per post, under each post's own
    name — duplication on disk is cheaper than a filename that only makes sense
    with the workbook open.

    `missing_asset_ids` reflects only the ORIGINAL file (storage_path). A missing
    or unreadable CONFORMED copy (publish_path) is noted in `problems` but never
    added there — the irreplaceable source still exported fine.

    Orphaned assets — rows in `assets` reached by no post, e.g. an abandoned
    compose upload, or the leftover original of a post that was deleted (assets
    is ON DELETE RESTRICT while post_assets cascades) — are backed up too, into
    UNLINKED_DIR, so a backup never silently drops content.
    """
    result = CopyResult()
    linked_asset_ids: set[int] = set()
    for post in bundle.posts:
        for image in post.images:
            linked_asset_ids.add(image.asset_id)
            _copy_one(
                _resolve(asset_root, image.storage_path),
                out_dir / IMAGES_DIR,
                image.export_filename,
                result,
                image.asset_id,
                is_original=True,
            )
            if image.publish_path and image.published_filename:
                _copy_one(
                    _resolve(asset_root, image.publish_path),
                    out_dir / PUBLISHED_DIR,
                    image.published_filename,
                    result,
                    image.asset_id,
                    is_original=False,
                )

    for asset in bundle.assets:
        if asset.asset_id in linked_asset_ids:
            continue
        _copy_one(
            _resolve(asset_root, asset.storage_path),
            out_dir / UNLINKED_DIR,
            unlinked_filename(asset),
            result,
            asset.asset_id,
            is_original=True,
        )
    return result


# Bump when the JSON shape changes incompatibly, so a future importer can branch.
#
# NOT bumped when the auto-fill columns moved to an "autofill_lanes" key: worker/restore.py
# is the only reader, it REFUSES any version it does not recognize, and it reads only the
# posts/images/assets fields — none of which changed. Bumping would have made every
# existing backup un-restorable to buy a distinction nothing consumes.
JSON_FORMAT_VERSION = 1


def write_json(bundle: ExportBundle, out_dir: Path) -> Path:
    """Full-fidelity machine-readable dump, for a future re-import.

    Nested rather than flat: a post CONTAINS its images, which a spreadsheet cannot
    express. Secrets are absent because collect.py never read them — this is not a
    raw table dump.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "format_version": JSON_FORMAT_VERSION,
        "generated_at": bundle.generated_at,
        "posts": [asdict(p) for p in bundle.posts],
        "sends": [asdict(s) for s in bundle.sends],
        "metrics": [asdict(m) for m in bundle.metrics],
        "assets": [asdict(a) for a in bundle.assets],
        "channels": [asdict(c) for c in bundle.channels],
        "channel_groups": [asdict(g) for g in bundle.channel_groups],
        "autofill_lanes": [asdict(lane) for lane in bundle.autofill_lanes],
    }
    path = out_dir / "export.json"
    # ensure_ascii=False keeps captions readable if someone opens this in a text editor.
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return path


WORKBOOK_NAME = "SocialScheduler-Export.xlsx"
MAX_COLUMN_WIDTH = 60
MAX_PROBLEM_LINES = 20


def _join(values: list[str] | list[int]) -> str:
    return ", ".join(str(v) for v in values)


def _sanitize_cell(value):
    """Strip control characters openpyxl refuses to write (IllegalCharacterError).

    Captions pasted from Word or a PDF, and API `last_error` strings, routinely
    carry \\x0b/\\x0c and friends. Only strings are touched — export.json keeps the
    original, unmodified text; this sanitization is workbook-only.
    """
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def _add_sheet(book: Workbook, title: str, headers: list[str], rows: list[list]) -> None:
    """One tab: a bold frozen header row, the data, and readable column widths."""
    sheet = book.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for row in rows:
        sheet.append([_sanitize_cell(v) for v in row])
    for index, header in enumerate(headers, start=1):
        widths = [len(str(header))]
        widths += [
            len(str(row[index - 1])) for row in rows if row[index - 1] is not None
        ]
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(widths) + 2, MAX_COLUMN_WIDTH
        )
    sheet.auto_filter.ref = sheet.dimensions


POSTS_HEADERS = [
    "post_id", "caption", "first_comment", "post_type", "content_kind", "content_status",
    "status", "is_bpp", "bpp_marked_at", "tags", "green_periods", "blackout_periods",
    "cooldown_days",
    "target_channels", "image_files", "times_posted", "last_posted_at_utc", "total_reach",
    "total_likes", "created_by", "created_at_utc",
]

SENDS_HEADERS = [
    "publication_id", "post_id", "caption_preview", "channel", "scheduled_at_local",
    "scheduled_at_utc", "published_at_local", "published_at_utc", "status", "is_held",
    "is_dry_run", "attempt_count", "last_error", "remote_post_id",
]

METRICS_HEADERS = [
    "publication_id", "post_id", "fetched_at_utc", "reach", "impressions", "likes",
    "comments", "saves", "shares", "video_views",
]

ASSETS_HEADERS = [
    "asset_id", "exported_filename", "original_filename", "media_kind", "width", "height",
    "byte_size", "conform_mode", "needs_review", "content_hash", "used_by_posts",
    "published_copy_filename",
]

# The auto-fill columns are absent from both of these on purpose: migration 0028 froze
# them and moved the real config to autofill_lanes, so printing them stated a cadence the
# worker was not running. They live on the "Auto-fill lanes" tab now.
CHANNELS_HEADERS = [
    "channel_id", "platform", "account_name", "business_label", "timezone", "is_active",
    "requires_approval", "remote_account_id", "linked_page_id", "group_id",
]

GROUPS_HEADERS = [
    "group_id", "name", "timezone", "is_active",
]

LANES_HEADERS = [
    "lane_id", "owner", "surface", "enabled", "cadence_config", "min_queue_depth",
    "target_queue_depth", "reuse_min_age_days",
]


def write_workbook(
    bundle: ExportBundle, out_dir: Path, missing_asset_ids: set[int]
) -> Path:
    """The human artifact: seven tabs, each with one clear grain.

    Written after the images are copied, because the Assets tab reports which files
    were missing from disk.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    book = Workbook()
    book.remove(book.active)  # drop openpyxl's default empty sheet

    _add_sheet(book, "Posts", POSTS_HEADERS, [
        [
            p.post_id, p.caption, p.first_comment, p.post_type, p.content_kind,
            p.content_status, p.status,
            # Written as YES/blank rather than TRUE/FALSE: this column is scanned by eye
            # in a spreadsheet, and a column of FALSE is harder to read past than blanks.
            "YES" if p.is_bpp else "", p.bpp_marked_at,
            _join(p.tags), _join(p.green_periods),
            _join(p.blackout_periods), p.cooldown_days, _join(p.target_channels),
            _join([i.export_filename for i in p.images]), p.times_posted,
            p.last_posted_at, p.total_reach, p.total_likes, p.created_by, p.created_at,
        ]
        for p in bundle.posts
    ])

    _add_sheet(book, "Sends", SENDS_HEADERS, [
        [
            s.publication_id, s.post_id, s.caption_preview, s.channel_label,
            s.scheduled_at_local, s.scheduled_at_utc, s.published_at_local,
            s.published_at_utc, s.status, s.is_held, s.is_dry_run, s.attempt_count,
            s.last_error, s.remote_post_id,
        ]
        for s in bundle.sends
    ])

    _add_sheet(book, "Metrics", METRICS_HEADERS, [
        [
            m.publication_id, m.post_id, m.fetched_at, m.reach, m.impressions, m.likes,
            m.comments, m.saves, m.shares, m.video_views,
        ]
        for m in bundle.metrics
    ])

    # An asset can appear in several posts under a different name in each.
    # `published` is index-aligned with `usage`/`names` (one entry per post, in the
    # same order) rather than a compacted list of just the conformed ones — otherwise
    # a reader lining these columns up positionally would pair the wrong post with
    # the wrong conformed file whenever only some of an asset's posts were conformed.
    usage: dict[int, list[int]] = {}
    names: dict[int, list[str]] = {}
    published: dict[int, list[str]] = {}
    for post in bundle.posts:
        for image in post.images:
            usage.setdefault(image.asset_id, []).append(post.post_id)
            names.setdefault(image.asset_id, []).append(image.export_filename)
            # Placeholder "-" holds this post's slot so the list stays index-aligned
            # with used_by_posts even when this particular image wasn't conformed.
            # Do NOT drop entries here to "clean up" the column — that's what
            # caused the misalignment bug this comment is guarding against.
            published.setdefault(image.asset_id, []).append(
                image.published_filename or "-"
            )

    def _published_cell(asset_id: int) -> str:
        entries = published.get(asset_id, [])
        # Nothing was conformed for this asset anywhere — an all-placeholder row
        # ("-, -, -") is just noise, so leave the cell empty instead.
        if not any(entry != "-" for entry in entries):
            return ""
        return _join(entries)

    def _exported_filename_cell(asset: ExportedAsset) -> str:
        if asset.asset_id in missing_asset_ids:
            return "MISSING"
        if asset.asset_id in names:
            return _join(names[asset.asset_id])
        # Not linked to any post — the file backed up under UNLINKED_DIR.
        return unlinked_filename(asset)

    _add_sheet(book, "Assets", ASSETS_HEADERS, [
        [
            a.asset_id,
            _exported_filename_cell(a),
            a.original_filename, a.media_kind, a.width, a.height, a.byte_size,
            a.conform_mode, a.needs_review, a.content_hash,
            _join(usage.get(a.asset_id, [])), _published_cell(a.asset_id),
        ]
        for a in bundle.assets
    ])

    _add_sheet(book, "Channels", CHANNELS_HEADERS, [
        [
            c.channel_id, c.platform, c.account_name, c.business_label, c.timezone,
            c.is_active, c.requires_approval,
            c.remote_account_id, c.linked_page_id, c.group_id,
        ]
        for c in bundle.channels
    ])

    # Its own sheet rather than columns folded into Channels: group membership belongs to
    # the GROUP, and repeating it on every member would read as per-channel settings that
    # could be edited independently. They cannot.
    _add_sheet(book, "Channel groups", GROUPS_HEADERS, [
        [g.group_id, g.name, g.timezone, g.is_active]
        for g in bundle.channel_groups
    ])

    # The tab that actually answers "what is auto-fill doing?". It cannot be folded into
    # Channels or Channel groups because an owner has one lane PER SURFACE — a group can
    # run a feed rotation and a Story rotation on independent cadences, and either tab
    # would have to pick one and hide the other. `owner` is denormalized to
    # "Name (channel|group)" so a row means something on its own.
    _add_sheet(book, "Auto-fill lanes", LANES_HEADERS, [
        [
            lane.lane_id, f"{lane.owner_name} ({lane.owner_kind})", lane.surface,
            lane.enabled, lane.cadence_config, lane.min_queue_depth,
            lane.target_queue_depth, lane.reuse_min_age_days,
        ]
        for lane in bundle.autofill_lanes
    ])

    # Captions are long; wrapping keeps the Posts tab scannable.
    for cell in book["Posts"]["B"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    path = out_dir / WORKBOOK_NAME
    book.save(path)
    return path


def write_readme(bundle: ExportBundle, out_dir: Path, copy_result: CopyResult) -> Path:
    """Plain-English orientation, for someone opening this folder years from now
    with no idea what produced it."""
    out_dir.mkdir(parents=True, exist_ok=True)

    linked_asset_ids = {img.asset_id for post in bundle.posts for img in post.images}
    has_unlinked_assets = any(a.asset_id not in linked_asset_ids for a in bundle.assets)

    lines = [
        "SocialScheduler — Content Export",
        "=" * 40,
        "",
        f"Created: {bundle.generated_at}",
        "",
        "WHAT'S IN HERE",
        "",
        "  socialscheduler.db            The restore file. This is the whole install —",
        "                                every post, schedule, tag and statistic — in the",
        "                                form the app itself reads. You never open this",
        "                                one by hand; it is what 'Restore-Mac' puts back",
        "                                if you ever need to rebuild on a new computer.",
        "                                Without it, the files below are a record you can",
        "                                read but not reload.",
        "",
        "  SocialScheduler-Export.xlsx   Open this one. Seven tabs: Posts, Sends,",
        "                                Metrics, Assets, Channels, Channel groups",
        "                                and Auto-fill lanes.",
        "                                Start on 'Posts' — it has everything you",
        "                                normally want, including how often each",
        "                                post has gone out and how it performed.",
        "                                'Auto-fill lanes' is where the posting",
        "                                schedules live: one row per account (or",
        "                                group) per place it posts, so a feed",
        "                                rotation and a Stories rotation each show",
        "                                their own cadence.",
        "",
        "  export.json                   The same data in a form software can read",
        "                                back in. You don't need to open this.",
        "",
        f"  {IMAGES_DIR}/                        Your original uploaded images.",
        "",
        f"  {PUBLISHED_DIR}/              The cropped/padded versions that were",
        "                                actually sent to Instagram, where those",
        "                                differ from the original.",
        "",
    ]
    if has_unlinked_assets:
        lines += [
            f"  {UNLINKED_DIR}/               Images not attached to any post — for",
            "                                example, left over from a deleted post or",
            "                                an abandoned upload that was never finished.",
            "                                They're included here so nothing is lost.",
            "",
        ]
    lines += [
        "Image filenames are 'postID_caption_position', so you can match any image",
        "back to its row in the Posts tab.",
        "",
        "WHAT IT CONTAINS",
        "",
        f"  {len(bundle.posts)} post(s)",
        f"  {len(bundle.sends)} scheduled or completed send(s)",
        f"  {len(bundle.metrics)} metric snapshot(s)",
        f"  {len(bundle.assets)} asset(s) on record",
        f"  {len(bundle.channels)} channel(s)",
        f"  {len(bundle.channel_groups)} channel group(s)",
        f"  {len(bundle.autofill_lanes)} auto-fill lane(s)",
        f"  {copy_result.copied} image file(s) copied",
        "",
        "HOW TO RESTORE THIS BACKUP",
        "",
        "  On a working install of SocialScheduler, double-click 'Restore-Mac' in the",
        "  app folder and choose THIS folder when it asks. It puts the database back",
        "  and rebuilds the image store from the images in here, renaming them back to",
        "  the names the app expects. Nothing is overwritten until you confirm, and",
        "  your current database is saved aside first.",
        "",
        "  Stop the app before restoring. The restore refuses to run while it is on.",
        "",
        "A NOTE ON SECURITY",
        "",
        "  Access credentials are deliberately NOT included in this export. If you",
        "  ever restore from it, you'll reconnect the accounts fresh. That's normal",
        "  — it means this folder is safe to store in Google Drive or Dropbox.",
        "",
    ]

    if not copy_result.problems:
        lines += ["PROBLEMS", "", "  No problems. Everything exported.", ""]
    else:
        lines += ["PROBLEMS", ""]
        if copy_result.missing_asset_ids:
            lines += [
                f"  {len(copy_result.missing_asset_ids)} image file(s) could not be found on",
                "  disk. Their rows are marked MISSING in the Assets tab.",
                "",
            ]
        else:
            lines += [
                "  Every original image exported successfully — nothing irreplaceable",
                "  was lost.",
                "",
            ]
        lines += [
            f"  {len(copy_result.problems)} problem(s) were recorded during export (this is",
            "  a count of failed copy operations, not images — one image used in",
            "  several posts can fail more than once, so this number may be larger",
            "  than the image count above).",
            "",
        ]
        shown = copy_result.problems[:MAX_PROBLEM_LINES]
        lines += [f"    - {p}" for p in shown]
        omitted = len(copy_result.problems) - len(shown)
        if omitted > 0:
            lines.append(f"    ... and {omitted} more problem(s) not shown here.")
        lines.append("")

    path = out_dir / "README.txt"
    path.write_text("\n".join(lines), encoding="utf-8")
    return path
